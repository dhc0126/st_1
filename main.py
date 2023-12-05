# -*- coding: utf-8 -*-
"""
Created on Tue Nov 22 13:02:13 2022

@author: Administrator
"""
import sqlite3

import math
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import streamlit as st
import streamlit_echarts as ste
# import plotly.figure_factory as ff
from openpyxl import load_workbook
from pyecharts import options as opts
from pyecharts.charts import Line
from pyecharts.charts import Pie
from pyecharts.charts import Radar
from pyecharts.globals import ThemeType

sns.set_style("darkgrid", {"axes.facecolor": ".9"})  # 设置直方图样式
plt.rcParams['axes.unicode_minus'] = False  # 正常显示负号
plt.rcParams['font.sans-serif'] = ['SimHei']  # 设置图标支持中文

# 首先获得颜色迭代Iterator对象:
colors = list(mcolors.TABLEAU_COLORS)  # 构建颜色列表
cnames = {
    'aliceblue': '#F0F8FF',
    # 'antiquewhite': '#FAEBD7',
    'aqua': '#00FFFF',
    # 'aquamarine': '#7FFFD4',
    # 'azure': '#F0FFFF',
    # 'beige': '#F5F5DC',
    # 'bisque': '#FFE4C4',
    # 'black': '#000000',
    # 'blanchedalmond': '#FFEBCD',
    # 'blue': '#0000FF',
    'blueviolet': '#8A2BE2',
    'brown': '#A52A2A',
    'burlywood': '#DEB887',
    'cadetblue': '#5F9EA0',
    'chartreuse': '#7FFF00',
    'chocolate': '#D2691E',
    'coral': '#FF7F50',
    'cornflowerblue': '#6495ED',
    'cornsilk': '#FFF8DC',
    'crimson': '#DC143C',
    'cyan': '#00FFFF',
    'darkblue': '#00008B',
    'darkcyan': '#008B8B',
    'darkgoldenrod': '#B8860B',
    'darkgray': '#A9A9A9',
    'darkgreen': '#006400',
    'darkkhaki': '#BDB76B',
    'darkmagenta': '#8B008B',
    'darkolivegreen': '#556B2F',
    'darkorange': '#FF8C00',
    'darkorchid': '#9932CC',
    'darkred': '#8B0000',
    'darksalmon': '#E9967A',
    'darkseagreen': '#8FBC8F',
    'darkslateblue': '#483D8B',
    'darkslategray': '#2F4F4F',
    'darkturquoise': '#00CED1',
    'darkviolet': '#9400D3',
    'deeppink': '#FF1493',
    'deepskyblue': '#00BFFF',
    'dimgray': '#696969',
    'dodgerblue': '#1E90FF',
    'firebrick': '#B22222',
    'floralwhite': '#FFFAF0',
    'forestgreen': '#228B22',
    'fuchsia': '#FF00FF',
    'gainsboro': '#DCDCDC',
    'ghostwhite': '#F8F8FF',
    'gold': '#FFD700',
    'goldenrod': '#DAA520',
    'gray': '#808080',
    'green': '#008000',
    'greenyellow': '#ADFF2F',
    'honeydew': '#F0FFF0',
    'hotpink': '#FF69B4',
    'indianred': '#CD5C5C',
    'indigo': '#4B0082',
    'ivory': '#FFFFF0',
    'khaki': '#F0E68C',
    'lavender': '#E6E6FA',
    'lavenderblush': '#FFF0F5',
    'lawngreen': '#7CFC00',
    'lemonchiffon': '#FFFACD',
    'lime': '#00FF00',
    'limegreen': '#32CD32',
    'linen': '#FAF0E6',
    'magenta': '#FF00FF',
    'maroon': '#800000',
    'mediumaquamarine': '#66CDAA',
    'mediumblue': '#0000CD',
    'mediumorchid': '#BA55D3',
    'mediumpurple': '#9370DB',
    'mediumseagreen': '#3CB371',
    'mediumslateblue': '#7B68EE',
    'mediumspringgreen': '#00FA9A',
    'mediumturquoise': '#48D1CC',
    'mediumvioletred': '#C71585',
    'midnightblue': '#191970',
    'mintcream': '#F5FFFA',
    'mistyrose': '#FFE4E1',
    'moccasin': '#FFE4B5',
    'navajowhite': '#FFDEAD',
    'navy': '#000080',
    'oldlace': '#FDF5E6',
    'olive': '#808000',
    'olivedrab': '#6B8E23',
    'orange': '#FFA500',
    'orangered': '#FF4500',
    'orchid': '#DA70D6',
    'palegoldenrod': '#EEE8AA',
    'palegreen': '#98FB98',
    'paleturquoise': '#AFEEEE',
    'palevioletred': '#DB7093',
    'papayawhip': '#FFEFD5',
    'peachpuff': '#FFDAB9',
    'peru': '#CD853F',
    'pink': '#FFC0CB',
    'plum': '#DDA0DD',
    'powderblue': '#B0E0E6',
    'purple': '#800080',
    'red': '#FF0000',
    'rosybrown': '#BC8F8F',
    'royalblue': '#4169E1',
    'saddlebrown': '#8B4513',
    'salmon': '#FA8072',
    'sandybrown': '#FAA460',
    'seagreen': '#2E8B57',
    'seashell': '#FFF5EE',
    'sienna': '#A0522D',
    'silver': '#C0C0C0',
    'skyblue': '#87CEEB',
    'slateblue': '#6A5ACD',
    'slategray': '#708090',
    'springgreen': '#00FF7F',
    'steelblue': '#4682B4',
    'tan': '#D2B48C',
    'teal': '#008080',
    'thistle': '#D8BFD8',
    'tomato': '#FF6347',
    'turquoise': '#40E0D0',
    'violet': '#EE82EE',
    'wheat': '#F5DEB3'}
list_color = [i for i in cnames.values()]


def pyech_pie(title,x_list,y_list,width="400px",height="400px"):
    c=(
        Pie(init_opts=opts.InitOpts(width="800px", height="400px"))
            .add("",
                 data_pair=[list(z) for z in zip(x_list, y_list)],
                 radius=["0", "50%"],
                 center=["50%", "50%"],
                 label_opts=opts.LabelOpts(is_show=True, formatter="{b}: {c} ({d}%)"),
                 )  ####展示数值
            # .set_colors(["blue", "green", "yellow", "red", "pink", "orange", "purple"])  ####设置扇形颜色
            .set_global_opts(
            title_opts=opts.TitleOpts(
                title=title,
                # subtitle=subtitle1,
                pos_left="center",
                pos_top="20",
                title_textstyle_opts=opts.TextStyleOpts(color="#fff"),  ####设置标题颜色
            ),
            legend_opts=opts.LegendOpts(is_show=True),
        )
            .set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{a} <br/>{b}: {c} ({d}%)"
            ),
        )

    )
    return c

# 1、成绩导入数据库
def td(f, sheet):
    writer = pd.read_excel(f, sheet_name=sheet, header=1)
    writer.fillna(0, inplace=True)
    # 设定成绩相关参数：科目，时间等
    # sublist = ('地理',"")

    sublist = writer.columns[1:]
    subject = st.selectbox(
        '科目', sublist[1:-2])

    con = sqlite3.connect("Test.db")
    cur = con.cursor()
    data = pd.read_sql_query('select * from %s where 班级 = %d ' % (subject, eval(sheet)), con)

    # 按考试日期排序
    colu=list(data.columns[2:])
    col=['班级','姓名']
    for i in range(len(colu)):
        a=max(colu)
        col.append(a)
        colu.remove(a)
    st.write('数据库中已有的成绩数据')
    st.write(data[col])


    cur.close()
    con.close()

    df = writer[['姓名', subject]]
    date_test = st.date_input(
        '本次测试日期')
    ddt = "d"
    for i in str(date_test):
        if i != '-':
            ddt = ddt + i

    st.write(date_test)
    # 导入数据库
    if st.button('导入到数据库'):
        st.write('Why hello there')

        con = sqlite3.connect("Test.db")
        cur = con.cursor()
        # cur.execute('ALTER TABLE '+subject+' ADD COLUMN '+ ddt +' integer ')
        try:
            sql = 'ALTER TABLE %s ADD COLUMN %s INTEGER '
            cur.execute(sql % (subject, ddt))
        except:
            st.write('字段已存在')
        con.commit()

        # 对于每一行，通过列名name访问对应的元素

        for index, row in df.iterrows():
            try:
                cur.execute("insert  into " + subject + " (班级,姓名,%s) values(%s,'%s',%d);" % (ddt, sheet, row['姓名'], row[subject]))
            except:
                cur.execute("UPDATE %s SET %s = %s WHERE 姓名 like '%s';" % (subject, ddt, row[subject], row['姓名']))

        con.commit()
        data = pd.read_sql_query('select * from ' + subject, con)
        st.write(data)
        cur.close()
        con.close()


    else:
        st.write('一旦导入数据将无法删除，请确认各信息无误后点击按钮')

    return


# 2、历次成绩对比
def lccjdb():
    con = sqlite3.connect("Test.db")
    cur = con.cursor()

    # sublist = ['地理','语文']
    sublist1 = cur.execute("SELECT name FROM sqlite_master WHERE type ='table'")
    sublist=[]
    for name in sublist1:
        sublist.append(name[0])

    subject = st.selectbox(
        '科目', sublist[1:])

    data = pd.read_sql_query('select * from '+ subject, con)

    del data['姓名']
    data = data.groupby('班级').mean()
    data = data.transpose()
    col1,col2=st.columns([1,2])

    with col1:
        st.write("#### 历次考试班级均分")
        data.sort_index(inplace=True)
        st.dataframe(data)
    with col2:
        st.write("#### 班级对比图")
        st.line_chart(data)

    st.write("### 班内对比图")
    clas = st.selectbox(
        '选择班级', data.columns)
    sql = "select 姓名 from 地理 where 班级 = %d" % (int(clas))
    stuname = pd.read_sql_query(sql, con)
    stu = st.multiselect('选出要对比的学生', stuname)

    data1 = data.sort_index()
    line = Line()

    line.add_xaxis(xaxis_data=data1.index.tolist())
    sql = "select * from 地理 where 班级 = %d" % (int(clas))
    dfstu = pd.read_sql_query(sql, con)

    dfstu = dfstu.drop(columns='班级')

    for i in stu:
        # st.write(i,dfstu[dfstu['姓名']==i])
        d = dfstu[dfstu['姓名'] == i].transpose()
        d = d.sort_index()
        # st.write(d[:-1].values)
        line.add_yaxis(series_name=i, y_axis=d[:-1].values.tolist())

    # ste.st_pyecharts(line)
    g = line.render_embed()
    st.components.v1.html(g, height=500)

    data2 = pd.DataFrame()
    for i in stu:
        d = dfstu[dfstu['姓名'] == i]
        d = d.sort_index()
        d = d.set_index(keys=['姓名'])
        # data2=data2.append(d)
        data2 = pd.concat([data2, d], ignore_index=True)
    data0 = data2.transpose()
    data0=data0.sort_index()
    data4=data0.sub(data[clas], axis=0)
    st.write("### 与均分差，进退步程度")
    st.line_chart(data4)


    cur.close()
    con.close()

    return


# 3、常规统计
def cgtj():
    '读取数据库'
    con = sqlite3.connect("Test.db")
    cur = con.cursor()

    sublist1 = cur.execute("SELECT name FROM sqlite_master WHERE type ='table'")
    sublist = []
    for name in sublist1:
        sublist.append(name[0])

    subject = st.selectbox(
        '科目', sublist[1:])

    data = pd.read_sql_query('select * from ' + subject, con)


    colu = list(data.columns[2:])
    col = []
    for i in range(len(colu)):
        a = max(colu)
        col.append(a)
        colu.remove(a)

    test = st.selectbox(
        '选择考试场次', col)

    df = data[['班级', test]]
    df = df.dropna()  # 清洗空值，很重要

    c = df.groupby('班级').mean()
    clist = ['全年级']
    for i in list(c.index):
        clist.append(str(i))
    s_c = st.selectbox('选出班级', clist)

    if s_c == '全年级':
        df = df
    else:
        df = df[df['班级'] == eval(s_c)]

    # 计算统计结果，出图

    col1, col2 = st.columns([3,1])
    with col2:
        st.write(s_c,'成绩基本统计量')
        st.dataframe(df[test].describe())
    with col1:

        bins=[0,40,60,80,90,101]
        seg=pd.cut(df[test],bins,right=False)

        counts=df.value_counts(seg,sort=False)


        x_list = ['低分<40', '不及格<60','及格<80','良好<90','优秀']
        y_list = counts
        c = pyech_pie(s_c+"各分数段占比",x_list, y_list)

        # ste.st_pyecharts(c,height=500)
        g = c.render_embed()
        st.components.v1.html(g, height=700, width=900)
    st.dataframe(data[['班级','姓名',test]])
    return


# 4、单科试卷得分明细分析
def dkmxfx(f):

    writer, school, s_c,s,c, zsd = readfile(f)

    # 全区各校成绩直方图
    st.header('全区各校成绩直方图')

    st.write(school)
    if school == 'z尖草坪区':
        st.write(s)
        st.write(s.index)
        qschool=st.selectbox( '学校', s.index)

    dt = writer[writer.学校 == school][writer.班级 == s_c]
    scroe = st.selectbox(
        '统计项目', ('总分', '主观分', '客观分'))
    z = dt[['姓名', scroe]]  # 按 “学校” 分组统计
    bins = np.arange(-10, 110, 12)
    g = sns.FacetGrid(z, height=3)
    g.map(sns.distplot, scroe, bins=bins)
    g.set_titles(col_template=school + s_c)  # 更改子图标题为：‘学校’+‘班级’
    g.set_ylabels(scroe + "频数")
    col1, col2 = st.columns(2)
    with col1:
        st.write(school + s_c + scroe)
        st.pyplot(g)
    with col2:
        st.write(z)


    # 按班级计算每小题均分.表格删除准考证号
    st.header('按班级计算每小题均分')

    dt = c.loc[s_c]
    d = dt.drop(labels=['总分', '主观分', '客观分'])
    cc = writer[writer['学校'] == 'z尖草坪区'].groupby('班级').mean()
    c = c.append(cc)
    st.bar_chart(d)
    st.dataframe(c)



    # 知识点得分情况
    st.header("知识点得分率")

    dtt = writer[writer['学校'] == school].groupby('班级').mean()
    dt = pd.DataFrame()
    ylim = []
    for i in zsd.columns:  # 迭代每个知识点

        a = 0
        for ii in zsd[i].values:  # 迭代知识点下的每个题目
            if a == 0:
                dt[i] = dtt[ii]
                y = writer[ii].max()
                a = a + 1
            else:
                try:
                    dt[i] = dtt[ii] + dt[i]
                    y = writer[ii].max() + y
                    y = math.ceil(y)
                    a = a + 1
                except:
                    continue
        ylim.append(y)

    ddd = dt / ylim * 100
    w = ddd.round(decimals=2)  # round(decimals=2)将平均数保留2位小数

    radar = Radar(init_opts=opts.InitOpts(theme=ThemeType.DARK))
    radar.set_series_opts(label_opts=opts.LabelOpts(is_show=False))
    radar.set_global_opts(legend_opts=opts.LegendOpts(type_='scroll'))

    # 将columns设为极坐标横轴
    c_schema = []
    for i in w.columns:
        c_schema.append({"name": i, "max": 100})
    radar.add_schema(schema=c_schema)

    n = 0
    for i in w.index:
        print(i)

        data = []
        for ii in w.loc[i, :]:
            data.append(ii)
        print(data)
        radar.add(series_name=i,
                  data=[data],
                  linestyle_opts=opts.LineStyleOpts(width=2, color=list_color[n]),
                  areastyle_opts=opts.AreaStyleOpts(opacity=0.5, color=list_color[n]))



        n = n + 1
    ste.st_pyecharts(radar,height=500)
    g = radar.render_embed()
    # st.components.v1.html(g, height=700, width=900)

    st.dataframe(w)

    return

def pianqu(f):
    st.write(f.name,"片区明细分析")
    writer = pd.read_excel(f, sheet_name="题目分析--学校", header=1)
    d=writer.groupby(["学校"]).sum()
    st.dataframe(d,800,800,)
    return


def readfile(f):

    writer = pd.read_excel(f, sheet_name="单科试卷得分明细表", header=1)
    zsd = pd.read_excel(f, sheet_name='知识点题目分布', header=1)
    # 根据全区学生成绩计算区平均成绩
    for pare in writer.columns[4:]:  # 计算每一小题均分
        writer.loc['全区平均分', pare] = writer.loc[:'全区平均分', pare].mean()
    # 将均分添加到‘z尖草坪区’，‘区均分班’ 里的成绩
    writer.loc['全区平均分', '学校'] = 'z尖草坪区'
    writer.loc['全区平均分', '班级'] = '区均分'

    with st.sidebar:
        s = writer.groupby('学校', ).mean()
        school = st.selectbox(
                '选出学校', s.index)
        c = writer[writer['学校'] == school].groupby('班级').mean()
        s_c = st.selectbox('选出班级', c.index)
    return  writer,school,s_c,s,c,zsd

with st.sidebar:
    """
    ## 学生成绩分析系统
     by：dhc
    """
    f = st.file_uploader("选择成绩文件")

    option = st.selectbox(
        '选择功能:',
        ('1、成绩导入数据库', '2、历次成绩对比', \
         '3、常规统计', '4、单科试卷得分明细分析', '5、2023片区'))

st.write('You selected:', option)

if option == '1、成绩导入数据库':

    wb = load_workbook(f)
    sheets = wb.sheetnames
    sheet = st.selectbox(
        '选出表单', sheets)
    td(f, sheet)

elif option == '2、历次成绩对比':

    lccjdb()

elif option == '3、常规统计':

    cgtj()

elif option == '4、单科试卷得分明细分析':

    dkmxfx(f)
    
elif option == '5、2023片区':
    
    pianqu(f)


